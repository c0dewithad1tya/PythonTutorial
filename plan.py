import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Define the lesson plan data
data = {
    "Week": [
        "Week 1", "Week 1", "Week 1", "Week 1", "Week 1", "Week 1", "Week 1",
        "Week 2", "Week 2", "Week 2", "Week 2", "Week 2", "Week 2", "Week 2",
        "Week 3", "Week 3", "Week 3", "Week 4", "Week 4", "Week 4"
    ],
    "Day": [
        "Day 1", "Day 2", "Day 3", "Day 4", "Day 5", "Day 6", "Day 7",
        "Day 8", "Day 9", "Day 10", "Day 11", "Day 12", "Day 13", "Day 14",
        "Week 3 - Day 1", "Week 3 - Day 2", "Week 3 - Day 3", "Week 4 - Day 1", "Week 4 - Day 2", "Week 4 - Day 3"
    ],
    "Topic": [
        "Kafka Basics – Pub/Sub, Topics, Brokers, Partitions",
        "Kafka Producers & Consumers (Console & Java/Python)",
        "Kafka Internals – Offsets, Consumer Groups",
        "Kafka Configs & Reliability (acks, retries, delivery)",
        "Schema Registry & Avro (Optional intro)",
        "Kafka in Payment Use Case",
        "Recap + Mini Project",
        "Kafka Streams Introduction – DSL & Concepts",
        "Stream transformations: map, filter, join",
        "Stateful operations – aggregations, windows",
        "KTable vs KStream",
        "Interactive Queries & Materialized Views",
        "Error Handling, Serialization",
        "Kafka Streams Mini Project",
        "Flink Architecture, Jobs, Operators",
        "Stateful vs Stateless Operators",
        "Watermarks, Event Time vs Processing Time",
        "Complex Event Processing (CEP)",
        "Integration with Kafka",
        "Mini Project – Real-time Payment Orchestrator"
    ],
    "Hands-on": [
        "Spin up Kafka using Docker, create a topic",
        "Write a basic producer & consumer",
        "Simulate parallel consumers & offset tracking",
        "Tweak configs & test message delivery",
        "Produce/consume Avro messages",
        "Payment pub/sub simulation",
        "Simple stream topology",
        "Transform and enrich data",
        "Count transactions per window",
        "Build account-level summaries",
        "Query Kafka stream state",
        "Use custom serializers",
        "Running fraud check stream app",
        "Setup Flink, run basic pipelines",
        "Implement a keyed state tracker",
        "Process delayed events",
        "Pattern match a payment fraud",
        "Source/Sink with Kafka topics",
        "End-to-end flow with Flink"
    ]
}

df = pd.DataFrame(data)
excel_path = "Kafka_Flink_Streams_Learning_Plan.xlsx"

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, sheet_name='Sheet1')
    worksheet = writer.sheets['Sheet1']

    # Set column widths
    worksheet.column_dimensions['A'].width = 10
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 50
    worksheet.column_dimensions['D'].width = 50

    # Header formatting
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for cell in worksheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Freeze header row and set auto filter
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions

    # Center align all cells and add borders
    center_alignment = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    max_col = worksheet.max_column
    max_row = worksheet.max_row

    for row in worksheet.iter_rows(min_row=1, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.alignment = center_alignment
            cell.border = border

    # Light blue background for data cells
    data_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    for row in worksheet.iter_rows(min_row=2, max_col=max_col, max_row=max_row):
        for cell in row:
            cell.fill = data_fill
